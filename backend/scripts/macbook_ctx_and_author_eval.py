#!/usr/bin/env python3
"""
MacBook Model Context & Author Extraction Evaluator

Runs 3 experiments against Ollama models on the MacBook inference cluster:
  1. Pull reported context window from /api/show
  2. Needle-in-haystack probe at multiple prompt sizes
  3. Author extraction from a real study document

Usage:
    python backend/scripts/macbook_ctx_and_author_eval.py \\
      --models macbookmodelnames.csv \\
      --file path/to/study.md \\
      --ground-truth "T.F.X. Collins, R.L. Sprando, ..." \\
      --delay 60 \\
      --timeout 600 \\
      --output ctx_author_eval.xlsx

Environment Variables:
    MACBOOK_LLM_BASE_URL: Override the default base URL
"""

import argparse
import json
import re
import sys
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# =============================================================================
# Constants
# =============================================================================

DEFAULT_BASE_URL = "http://macbook1.sciencegpt.ca"
CONTEXT_THRESHOLD = 32768  # 32k tokens
PROBE_SIZES = [4096, 8192, 16384, 32768, 65536, 131072, 262144]  # characters

DEFAULT_GROUND_TRUTH = (
    "T.F.X. Collins, R.L. Sprando, T.N. Black, M.E. Shackelford, "
    "N. Olejnik, M.J. Ames, J.I. Rorie, D.I. Ruggles"
)

# Seed text for needle probe filler — scientific-sounding to avoid model refusals
_FILLER_SEED = (
    "The experimental design incorporated multiple cohorts of Sprague-Dawley rats "
    "assigned to treatment groups based on stratified randomization. "
    "Animals were housed in a controlled environment with a 12-hour light/dark cycle "
    "and provided ad libitum access to standard rodent chow and filtered drinking water. "
    "Body weights were recorded weekly throughout the study period. "
    "Hematological and biochemical parameters were assessed at scheduled necropsy intervals "
    "in accordance with established OECD testing guidelines for reproductive toxicology. "
    "Histopathological examination of target organs was performed by a board-certified "
    "veterinary pathologist blinded to treatment allocation. "
    "Statistical analysis employed a one-way ANOVA with Dunnett post-hoc correction "
    "for multiple comparisons against the vehicle control group. "
)

# Excel styling colours
COLOUR_HEADER = "D9D9D9"
COLOUR_GREEN = "C6EFCE"
COLOUR_RED = "FFC7CE"
COLOUR_SKIP = "FFEB9C"

# =============================================================================
# Model list (38 models, same order as evaluate_macbook_models.py)
# =============================================================================

EXPECTED_MODELS: List[str] = [
    # 3B - 4B
    "llama3.2:3b-instruct-fp16",
    "llama3.2:3b-instruct-q4_K_M",
    "MedAIBase/MedGemma1.5:4b",
    "phi4-mini:3.8b",
    "phi3.5:3.8b",
    "nemotron-mini:4b-instruct-q4_K_M",
    "nemotron-mini:4b-instruct-fp16",
    "nemotron-mini:4b-instruct-q8_0",
    # 7B - 8B
    "llama3.1:8b",
    "llama3.1:8b-instruct-q4_K_M",
    "dolphin-llama3:8b",
    "openbiollm-llama-3:8b-q8_0",
    "openbiollm-llama-3:8b-q6_k",
    "Mistral-7B-Instruct-v0.3-Q4_K_M:latest",
    "qwen3:8b-q4_K_M",
    "rnj-1:8b",
    # 12B - 14B
    "ministral-3:14b-instruct-2512-q4_K_M",
    "mistral-nemo:12b",
    "mistral-nemo:12b-instruct-2407-q3_K_M",
    "gemma3:12b",
    "qwen3:14b-q4_K_M",
    "phi4-reasoning:14b",
    # 20B - 24B
    "mistral-small3.1:24b",
    "gpt-o3:20b",
    # 27B - 32B
    "gemma3:27b-it-qat",
    "gemma3:27b-it-q4_K_M",
    "gemma2:27b",
    "qwen3:30b-a3b-q4_K_M",
    "qwen2.5:32b",
    "nemotron-3-nano:30b",
    "olmo-2:32b-think-q4_K_M",
    # 70B+
    "llama3.3:70b-instruct-q2_K",
    "llama3.1:70b-instruct-q2_k",
    "llama3.1:70b",
    "openbiollm-llama-3:70b_q4_k_m",
    "mistral-large:123b",
    "nemotron:70b-instruct-q3_K_M",
    "nemotron:70b-instruct-q2_K",
]


def load_models(csv_path: Optional[str]) -> List[str]:
    """Load model list from CSV or fall back to EXPECTED_MODELS.

    The CSV format used by this project concatenates model names within cells
    without separators, so we match against EXPECTED_MODELS by normalizing
    punctuation (same approach as ModelListParser in evaluate_macbook_models.py).
    """
    if csv_path and Path(csv_path).exists():
        try:
            csv_text = Path(csv_path).read_text(encoding="utf-8")

            def _norm(s: str) -> str:
                return s.lower().replace("-", "").replace(":", "").replace("_", "")

            csv_norm = _norm(csv_text)
            matched = [m for m in EXPECTED_MODELS if _norm(m) in csv_norm]
            if matched:
                print(f"Loaded {len(matched)} models from {csv_path}")
                return matched
        except Exception as exc:
            print(f"Warning: could not parse {csv_path}: {exc}")

    print(f"Using built-in EXPECTED_MODELS list ({len(EXPECTED_MODELS)} models)")
    return list(EXPECTED_MODELS)


# =============================================================================
# Config
# =============================================================================


@dataclass
class Config:
    base_url: str = DEFAULT_BASE_URL
    delay: int = 60  # seconds between generate calls
    timeout: int = 600  # request timeout in seconds
    output: str = ""  # set in main
    study_file: str = ""
    ground_truth: str = DEFAULT_GROUND_TRUTH
    max_tokens_probe: int = 256
    max_tokens_extract: int = 512


def parse_args() -> Tuple[Config, Optional[str]]:
    parser = argparse.ArgumentParser(
        description="MacBook Ollama model context & author extraction evaluator"
    )
    parser.add_argument("--models", default=None, help="Path to macbookmodelnames.csv")
    parser.add_argument(
        "--file",
        required=True,
        help="Path to the local .md study file for author extraction",
    )
    parser.add_argument(
        "--ground-truth",
        default=DEFAULT_GROUND_TRUTH,
        help="Expected author string for scoring",
    )
    parser.add_argument(
        "--delay",
        type=int,
        default=60,
        help="Seconds to wait between API generate calls (default: 60)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=600,
        help="Per-request timeout in seconds (default: 600)",
    )
    parser.add_argument(
        "--base-url",
        default=None,
        help=f"Ollama base URL (default: {DEFAULT_BASE_URL})",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel filename (default: ctx_author_eval_<timestamp>.xlsx)",
    )
    args = parser.parse_args()

    import os

    base_url = args.base_url or os.environ.get("MACBOOK_LLM_BASE_URL", DEFAULT_BASE_URL)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or f"ctx_author_eval_{timestamp}.xlsx"

    cfg = Config(
        base_url=base_url.rstrip("/"),
        delay=args.delay,
        timeout=args.timeout,
        output=output,
        study_file=args.file,
        ground_truth=args.ground_truth,
    )
    return cfg, args.models


# =============================================================================
# Data classes
# =============================================================================


@dataclass
class ContextResult:
    model: str
    ctx_int: Optional[int]  # None = unknown
    field_used: str
    status: str  # "pass" | "filtered" | "unknown"
    error: Optional[str]


@dataclass
class ProbeResult:
    model: str
    size_target: int
    size_actual: int
    key_a_ok: Optional[bool]  # None = skipped / error
    key_b_ok: Optional[bool]
    both_ok: Optional[bool]
    latency: Optional[float]
    skipped: bool = False
    error: Optional[str] = None
    response_snippet: str = ""


@dataclass
class AuthorResult:
    model: str
    response: str
    score_pct: float
    latency: Optional[float]
    error: Optional[str]


# =============================================================================
# Ollama API helpers
# =============================================================================


def fetch_context_window(
    session: requests.Session, base_url: str, model: str, timeout: int = 30
) -> Tuple[Optional[int], str, Optional[str]]:
    """Query /api/show and extract the model's reported context window.

    Returns (ctx_tokens, field_name_used, error_str).
    Mirrors the parse chain in the boss's bash script.
    """
    url = f"{base_url}/api/show"
    try:
        resp = session.post(url, json={"name": model}, timeout=timeout)
        if not resp.ok:
            return None, "", f"HTTP {resp.status_code}: {resp.text[:200]}"
        data = resp.json()
    except requests.exceptions.RequestException as exc:
        return None, "", str(exc)
    except json.JSONDecodeError as exc:
        return None, "", f"JSON decode error: {exc}"

    model_info = data.get("model_info", {})

    # 1. llama.context_length (most common for GGUF Llama-family models)
    val = model_info.get("llama.context_length")
    if val is not None:
        try:
            return int(val), "model_info.llama.context_length", None
        except (ValueError, TypeError):
            pass

    # 2. generic context_length key in model_info
    val = model_info.get("context_length")
    if val is not None:
        try:
            return int(val), "model_info.context_length", None
        except (ValueError, TypeError):
            pass

    # 3. Scan all model_info keys that contain "context"
    for key, v in model_info.items():
        if "context" in key.lower() and v is not None:
            try:
                return int(v), f"model_info.{key}", None
            except (ValueError, TypeError):
                pass

    # 4. parameters block (dict or string)
    params = data.get("parameters")
    if isinstance(params, dict):
        val = params.get("num_ctx")
        if val is not None:
            try:
                return int(val), "parameters.num_ctx (dict)", None
            except (ValueError, TypeError):
                pass
    elif isinstance(params, str):
        match = re.search(r"num_ctx\s*=?\s*(\d+)", params, re.IGNORECASE)
        if match:
            return int(match.group(1)), "parameters num_ctx (string)", None

    # 5. modelfile parameters section
    modelfile = data.get("modelfile", "")
    if modelfile:
        match = re.search(r"PARAMETER\s+num_ctx\s+(\d+)", modelfile, re.IGNORECASE)
        if match:
            return int(match.group(1)), "modelfile PARAMETER num_ctx", None

    return None, "", "no context field found in /api/show response"


def call_generate(
    session: requests.Session,
    base_url: str,
    model: str,
    prompt: str,
    timeout: int,
    max_tokens: int = 256,
    num_ctx: Optional[int] = None,
    temperature: float = 0.0,
) -> Tuple[str, float, Optional[str]]:
    """POST /api/generate and return (response_text, latency_s, error).

    Uses Ollama's options sub-object for generation parameters.
    """
    url = f"{base_url}/api/generate"

    options: Dict = {
        "temperature": temperature,
        "num_predict": max_tokens,
    }
    if num_ctx is not None:
        options["num_ctx"] = num_ctx

    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": options,
    }

    start = time.time()
    try:
        resp = session.post(url, json=payload, timeout=timeout)
        latency = time.time() - start

        if not resp.ok:
            return "", latency, f"HTTP {resp.status_code}: {resp.text[:200]}"

        data = resp.json()
        content = data.get("response") or data.get("content") or data.get("text", "")
        return content, latency, None

    except requests.exceptions.Timeout:
        return "", time.time() - start, f"Timeout after {timeout}s"
    except requests.exceptions.RequestException as exc:
        return "", time.time() - start, str(exc)
    except json.JSONDecodeError as exc:
        return "", time.time() - start, f"JSON decode error: {exc}"


# =============================================================================
# Experiment 1: Context reported
# =============================================================================


def run_experiment_1(
    models: List[str],
    cfg: Config,
    all_exp1: List[ContextResult],
    all_exp2: List[ProbeResult],
    all_exp3: List[AuthorResult],
) -> None:
    """Query /api/show for each model and record the reported context window.

    Appends to all_exp1 in-place and saves the Excel file after every result.
    Skips models already present in all_exp1 (resume support).
    """
    done = {r.model for r in all_exp1}
    total = len(models)

    print(f"\n{'='*60}")
    print(f"EXPERIMENT 1: Context window (reported by /api/show)")
    print(f"{'='*60}")

    session = requests.Session()
    try:
        for i, model in enumerate(models, 1):
            if model in done:
                print(f"  [{i:02d}/{total}] {model} ... RESUMED (already saved)")
                continue

            print(f"  [{i:02d}/{total}] {model} ...", end=" ", flush=True)
            ctx, field_used, error = fetch_context_window(
                session, cfg.base_url, model, timeout=30
            )

            if error and ctx is None:
                status = "unknown"
                print(f"unknown  ({error})")
            elif ctx is None:
                status = "unknown"
                print("unknown  (no field found)")
            elif ctx < CONTEXT_THRESHOLD:
                status = "filtered"
                print(f"FILTERED  ({ctx:,} < {CONTEXT_THRESHOLD:,}) [{field_used}]")
            else:
                status = "pass"
                print(f"pass  ({ctx:,}) [{field_used}]")

            all_exp1.append(
                ContextResult(
                    model=model,
                    ctx_int=ctx,
                    field_used=field_used,
                    status=status,
                    error=error,
                )
            )
            write_excel(cfg.output, all_exp1, all_exp2, all_exp3)
            time.sleep(0.5)
    finally:
        session.close()

    passed = sum(1 for r in all_exp1 if r.status == "pass")
    filtered = sum(1 for r in all_exp1 if r.status == "filtered")
    unknown = sum(1 for r in all_exp1 if r.status == "unknown")
    print(f"\n  Summary: {passed} pass, {filtered} filtered, {unknown} unknown")


# =============================================================================
# Experiment 2: Needle probe
# =============================================================================


def _build_filler(target_chars: int) -> str:
    """Generate filler text of approximately target_chars characters."""
    repeats = (target_chars // len(_FILLER_SEED)) + 2
    text = _FILLER_SEED * repeats
    return text[:target_chars]


def build_needle_prompt(target_size: int, key_a: str, key_b: str) -> str:
    """Build a prompt with key_a near the top, key_b near the bottom, filler between.

    Resulting prompt will be approximately target_size characters.
    """
    preamble = (
        "Instructions: Read the following passage carefully from start to finish. "
        "At the very end of your response return ONLY a JSON object — no other text.\n\n"
        f"RETRIEVAL_KEY_A: {key_a}\n\n"
    )
    epilogue = (
        f"\n\nRETRIEVAL_KEY_B: {key_b}\n\n"
        "Return ONLY this JSON (fill in the values you found above):\n"
        '{"key_a": "<value of RETRIEVAL_KEY_A>", "key_b": "<value of RETRIEVAL_KEY_B>"}'
    )

    filler_target = max(0, target_size - len(preamble) - len(epilogue))
    filler = _build_filler(filler_target)
    prompt = preamble + filler + epilogue
    return prompt


def _parse_probe_response(response: str, key_a: str, key_b: str) -> Tuple[bool, bool]:
    """Return (key_a_found, key_b_found) from model response.

    Tries JSON parse first, falls back to substring search.
    """
    # Clean up common model wrappers (```json ... ```)
    cleaned = re.sub(r"```(?:json)?", "", response).strip()
    # Find the last JSON-like object in the response
    json_match = re.search(r'\{[^{}]*"key_a"[^{}]*\}', cleaned, re.DOTALL)
    if json_match:
        try:
            parsed = json.loads(json_match.group())
            ka = str(parsed.get("key_a", "")).strip()
            kb = str(parsed.get("key_b", "")).strip()
            return ka == key_a, kb == key_b
        except json.JSONDecodeError:
            pass

    # Fallback: substring search
    ka_found = bool(re.search(re.escape(key_a), response))
    kb_found = bool(re.search(re.escape(key_b), response))
    return ka_found, kb_found


def run_experiment_2(
    models: List[str],
    ctx_results: Dict[str, ContextResult],
    cfg: Config,
    all_exp1: List[ContextResult],
    all_exp2: List[ProbeResult],
    all_exp3: List[AuthorResult],
) -> None:
    """Run needle-in-haystack probes at multiple prompt sizes.

    Appends to all_exp2 in-place and saves the Excel file after every result.
    Skips (model, size_target) pairs already present in all_exp2 (resume support).
    """
    done = {(r.model, r.size_target) for r in all_exp2}
    total_calls = sum(len(PROBE_SIZES) for _ in models)
    call_num = 0

    print(f"\n{'='*60}")
    print(
        f"EXPERIMENT 2: Needle probe ({len(models)} models × {len(PROBE_SIZES)} sizes)"
    )
    print(f"  Delay between calls: {cfg.delay}s")
    print(f"{'='*60}")

    session = requests.Session()
    try:
        for model in models:
            ctx_info = ctx_results.get(model)
            reported_ctx = ctx_info.ctx_int if ctx_info else None

            # Generate fresh unique keys for any sizes not yet done for this model
            key_a = uuid.uuid4().hex[:12]
            key_b = uuid.uuid4().hex[:12]

            print(f"\n  Model: {model}  (reported ctx: {reported_ctx or 'unknown'})")
            print(f"  Keys: KEY_A={key_a}  KEY_B={key_b}")

            for size_target in PROBE_SIZES:
                call_num += 1
                label = f"{size_target // 1024}k"

                if (model, size_target) in done:
                    print(
                        f"    [{call_num:03d}/{total_calls}] "
                        f"size={label:>5} → RESUMED (already saved)"
                    )
                    continue

                # Skip if clearly beyond this model's reported context
                # Use 4 chars/token as conservative estimate
                if reported_ctx is not None and size_target > reported_ctx * 4 * 1.1:
                    print(
                        f"    [{call_num:03d}/{total_calls}] "
                        f"size={label:>5} → SKIPPED (beyond reported ctx)"
                    )
                    all_exp2.append(
                        ProbeResult(
                            model=model,
                            size_target=size_target,
                            size_actual=0,
                            key_a_ok=None,
                            key_b_ok=None,
                            both_ok=None,
                            latency=None,
                            skipped=True,
                        )
                    )
                    write_excel(cfg.output, all_exp1, all_exp2, all_exp3)
                    continue

                prompt = build_needle_prompt(size_target, key_a, key_b)
                size_actual = len(prompt)

                # Set num_ctx to fit the prompt (rough token estimate + output buffer)
                num_ctx = max(size_actual // 3 + 512, 4096)

                print(
                    f"    [{call_num:03d}/{total_calls}] "
                    f"size={label:>5} ({size_actual:,} chars, ~{num_ctx:,} ctx tokens) ...",
                    end=" ",
                    flush=True,
                )

                response, latency, error = call_generate(
                    session,
                    cfg.base_url,
                    model,
                    prompt,
                    timeout=cfg.timeout,
                    max_tokens=cfg.max_tokens_probe,
                    num_ctx=num_ctx,
                )

                if error:
                    print(f"ERROR ({error[:60]})")
                    all_exp2.append(
                        ProbeResult(
                            model=model,
                            size_target=size_target,
                            size_actual=size_actual,
                            key_a_ok=None,
                            key_b_ok=None,
                            both_ok=None,
                            latency=latency,
                            error=error,
                            response_snippet=response[:300],
                        )
                    )
                else:
                    ka_ok, kb_ok = _parse_probe_response(response, key_a, key_b)
                    both_ok = ka_ok and kb_ok
                    symbol = "✓" if both_ok else ("~" if (ka_ok or kb_ok) else "✗")
                    print(
                        f"{symbol}  key_a={'Y' if ka_ok else 'N'}  "
                        f"key_b={'Y' if kb_ok else 'N'}  "
                        f"({latency:.1f}s)"
                    )
                    all_exp2.append(
                        ProbeResult(
                            model=model,
                            size_target=size_target,
                            size_actual=size_actual,
                            key_a_ok=ka_ok,
                            key_b_ok=kb_ok,
                            both_ok=both_ok,
                            latency=latency,
                            response_snippet=response[:300],
                        )
                    )

                write_excel(cfg.output, all_exp1, all_exp2, all_exp3)
                time.sleep(cfg.delay)
    finally:
        session.close()


# =============================================================================
# Experiment 3: Author extraction
# =============================================================================

_AUTHOR_PROMPT_TEMPLATE = """\
<markdown study>
{study_text}
</markdown study>

Extract the study author(s).

Few-shot examples:
Input: "The study was conducted by John Doe, Jane Smith, and their team."
Output: "John Doe, Jane Smith"

Input: "Authors: Maria Garcia, David Chen."
Output: "Maria Garcia, David Chen"

Extract the study author(s).

Output only the extracted information, nothing else."""


def score_authors(response: str, ground_truth: str) -> float:
    """Heuristic 0.0–1.0 score for author extraction.

    Extracts last names from ground_truth and checks how many appear
    in the model response. Partial credit for "et al." with correct first author.
    """
    if not response:
        return 0.0

    resp_lower = response.lower()

    def _last_name(token: str) -> str:
        """Extract last name from a token like 'T.F.X. Collins' → 'collins'."""
        parts = token.strip().replace(",", "").split()
        if not parts:
            return ""
        # Last part after any initials
        return parts[-1].lower()

    gt_tokens = [t.strip() for t in ground_truth.split(",") if t.strip()]
    last_names = [_last_name(t) for t in gt_tokens if _last_name(t)]

    if not last_names:
        return 0.0

    matched = sum(1 for ln in last_names if ln in resp_lower)
    base_score = matched / len(last_names)

    # "et al." with the first author correct → floor score at 0.5
    if "et al" in resp_lower and last_names[0] in resp_lower:
        return max(base_score, 0.5)

    return base_score


def run_experiment_3(
    models: List[str],
    study_text: str,
    cfg: Config,
    all_exp1: List[ContextResult],
    all_exp2: List[ProbeResult],
    all_exp3: List[AuthorResult],
) -> None:
    """Run author extraction on the study text for each model.

    Appends to all_exp3 in-place and saves the Excel file after every result.
    Skips models already present in all_exp3 (resume support).
    """
    done = {r.model for r in all_exp3}
    total = len(models)
    prompt = _AUTHOR_PROMPT_TEMPLATE.format(study_text=study_text)
    prompt_chars = len(prompt)
    # Token estimate for num_ctx: enough for the prompt + output
    num_ctx = max(prompt_chars // 3 + 1024, 8192)

    print(f"\n{'='*60}")
    print(f"EXPERIMENT 3: Author extraction ({total} models)")
    print(f"  Prompt size: {prompt_chars:,} chars (~{num_ctx:,} ctx tokens)")
    print(f"  Ground truth: {cfg.ground_truth[:80]}...")
    print(f"  Delay between calls: {cfg.delay}s")
    print(f"{'='*60}")

    session = requests.Session()
    try:
        for i, model in enumerate(models, 1):
            if model in done:
                print(f"  [{i:02d}/{total}] {model} ... RESUMED (already saved)")
                continue

            print(f"  [{i:02d}/{total}] {model} ...", end=" ", flush=True)

            response, latency, error = call_generate(
                session,
                cfg.base_url,
                model,
                prompt,
                timeout=cfg.timeout,
                max_tokens=cfg.max_tokens_extract,
                num_ctx=num_ctx,
            )

            if error:
                print(f"ERROR ({error[:60]})")
                all_exp3.append(
                    AuthorResult(
                        model=model,
                        response="",
                        score_pct=0.0,
                        latency=latency,
                        error=error,
                    )
                )
            else:
                score = score_authors(response, cfg.ground_truth)
                print(f"score={score:.0%}  ({latency:.1f}s)  → {response[:80]!r}")
                all_exp3.append(
                    AuthorResult(
                        model=model,
                        response=response,
                        score_pct=score * 100,
                        latency=latency,
                        error=None,
                    )
                )

            write_excel(cfg.output, all_exp1, all_exp2, all_exp3)
            time.sleep(cfg.delay)
    finally:
        session.close()


# =============================================================================
# Excel writer
# =============================================================================


def _header_style(cell: object) -> None:
    """Apply bold + gray fill + center alignment to a header cell."""
    cell.font = Font(bold=True)
    cell.fill = PatternFill(
        start_color=COLOUR_HEADER, end_color=COLOUR_HEADER, fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _bool_cell(
    ws, row: int, col: int, value: Optional[bool], skipped: bool = False
) -> None:
    """Write a boolean cell with colour coding."""
    cell = ws.cell(row=row, column=col)
    if skipped or value is None:
        cell.value = "skip" if skipped else "error"
        cell.fill = PatternFill(
            start_color=COLOUR_SKIP, end_color=COLOUR_SKIP, fill_type="solid"
        )
    elif value:
        cell.value = "Y"
        cell.fill = PatternFill(
            start_color=COLOUR_GREEN, end_color=COLOUR_GREEN, fill_type="solid"
        )
    else:
        cell.value = "N"
        cell.fill = PatternFill(
            start_color=COLOUR_RED, end_color=COLOUR_RED, fill_type="solid"
        )
    cell.alignment = Alignment(horizontal="center")


def _write_context_reported(wb: Workbook, results: List[ContextResult]) -> None:
    ws = wb.create_sheet("context_reported")
    headers = ["Model", "Reported num_ctx", "Field Used", "Status", "Error"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    status_colours = {
        "pass": COLOUR_GREEN,
        "filtered": COLOUR_RED,
        "unknown": COLOUR_SKIP,
    }

    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.model)
        ws.cell(row=row, column=2, value=r.ctx_int if r.ctx_int is not None else "")
        ws.cell(row=row, column=3, value=r.field_used)
        status_cell = ws.cell(row=row, column=4, value=r.status)
        colour = status_colours.get(r.status, COLOUR_HEADER)
        status_cell.fill = PatternFill(
            start_color=colour, end_color=colour, fill_type="solid"
        )
        status_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=r.error or "")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50
    ws.freeze_panes = "A2"


def _write_context_probe(wb: Workbook, results: List[ProbeResult]) -> None:
    ws = wb.create_sheet("context_probe")
    headers = [
        "Model",
        "Size Target (chars)",
        "Size Actual (chars)",
        "Key A Found",
        "Key B Found",
        "Both Found",
        "Latency (s)",
        "Error",
        "Response Snippet",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.model)
        ws.cell(row=row, column=2, value=r.size_target)
        ws.cell(row=row, column=3, value=r.size_actual if not r.skipped else "")
        _bool_cell(ws, row, 4, r.key_a_ok, r.skipped)
        _bool_cell(ws, row, 5, r.key_b_ok, r.skipped)
        _bool_cell(ws, row, 6, r.both_ok, r.skipped)
        ws.cell(row=row, column=7, value=f"{r.latency:.1f}" if r.latency else "")
        ws.cell(row=row, column=8, value=r.error or "")
        ws.cell(row=row, column=9, value=r.response_snippet)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 60
    ws.freeze_panes = "A2"


def _write_author_extraction(wb: Workbook, results: List[AuthorResult]) -> None:
    ws = wb.create_sheet("author_extraction")
    headers = ["Model", "Response", "Score (%)", "Latency (s)", "Error"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.model)
        ws.cell(row=row, column=2, value=r.response)
        score_cell = ws.cell(row=row, column=3, value=round(r.score_pct, 1))
        if r.error:
            score_cell.fill = PatternFill(
                start_color=COLOUR_RED, end_color=COLOUR_RED, fill_type="solid"
            )
        elif r.score_pct >= 80:
            score_cell.fill = PatternFill(
                start_color=COLOUR_GREEN, end_color=COLOUR_GREEN, fill_type="solid"
            )
        score_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=f"{r.latency:.1f}" if r.latency else "")
        ws.cell(row=row, column=5, value=r.error or "")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 40
    ws.freeze_panes = "A2"


def write_excel(
    output_path: str,
    exp1: List[ContextResult],
    exp2: List[ProbeResult],
    exp3: List[AuthorResult],
) -> None:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _write_context_reported(wb, exp1)
    _write_context_probe(wb, exp2)
    _write_author_extraction(wb, exp3)

    wb.save(output_path)
    print(f"\nExcel written → {output_path}")


# =============================================================================
# Resume: load existing results from a previous run's Excel file
# =============================================================================


def load_existing_results(
    output_path: str,
) -> Tuple[List[ContextResult], List[ProbeResult], List[AuthorResult]]:
    """Read back previously saved results from an existing Excel file.

    Returns empty lists if the file does not exist or cannot be read.
    """
    path = Path(output_path)
    if not path.exists():
        return [], [], []

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Warning: could not open {output_path} for resume: {exc}")
        return [], [], []

    def _safe_int(v) -> Optional[int]:
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    def _safe_float(v) -> Optional[float]:
        try:
            return float(str(v).strip())
        except (TypeError, ValueError):
            return None

    def _parse_bool_cell(v) -> Optional[bool]:
        if v == "Y":
            return True
        if v == "N":
            return False
        return None  # "skip", "error", blank

    # --- context_reported ---
    exp1: List[ContextResult] = []
    if "context_reported" in wb.sheetnames:
        ws = wb["context_reported"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            model, ctx_val, field_used, status, error = (list(row) + [None] * 5)[:5]
            exp1.append(
                ContextResult(
                    model=str(model),
                    ctx_int=_safe_int(ctx_val),
                    field_used=str(field_used or ""),
                    status=str(status or "unknown"),
                    error=str(error) if error else None,
                )
            )

    # --- context_probe ---
    exp2: List[ProbeResult] = []
    if "context_probe" in wb.sheetnames:
        ws = wb["context_probe"]
        for row in wb["context_probe"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cols = (list(row) + [None] * 9)[:9]
            model, size_target, size_actual, ka, kb, both, latency, error, snippet = (
                cols
            )
            skipped = str(ka).strip().lower() == "skip"
            exp2.append(
                ProbeResult(
                    model=str(model),
                    size_target=_safe_int(size_target) or 0,
                    size_actual=_safe_int(size_actual) or 0,
                    key_a_ok=_parse_bool_cell(ka),
                    key_b_ok=_parse_bool_cell(kb),
                    both_ok=_parse_bool_cell(both),
                    latency=_safe_float(latency),
                    skipped=skipped,
                    error=str(error) if error else None,
                    response_snippet=str(snippet or ""),
                )
            )

    # --- author_extraction ---
    exp3: List[AuthorResult] = []
    if "author_extraction" in wb.sheetnames:
        for row in wb["author_extraction"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cols = (list(row) + [None] * 5)[:5]
            model, response, score_pct, latency, error = cols
            exp3.append(
                AuthorResult(
                    model=str(model),
                    response=str(response or ""),
                    score_pct=_safe_float(score_pct) or 0.0,
                    latency=_safe_float(latency),
                    error=str(error) if error else None,
                )
            )

    wb.close()
    return exp1, exp2, exp3


# =============================================================================
# Helpers: filtering and family classification
# =============================================================================


def classify_family(model_name: str) -> str:
    """Map a model name to a broad family bucket."""
    name = model_name.lower()
    if re.search(r"gemma", name):
        return "gemma"
    if re.search(r"phi", name):
        return "phi"
    if re.search(r"llama|openbiollm|dolphin", name):
        return "llama"
    if re.search(r"mistral|ministral", name):
        return "mistral"
    if re.search(r"qwen", name):
        return "qwen"
    if re.search(r"nemotron", name):
        return "nemotron"
    return "other"


def _probe_passed_64k(model: str, probe_results: List[ProbeResult]) -> bool:
    """Return True if this model passed the both-keys test at the 64k char probe."""
    target = 65536
    for r in probe_results:
        if r.model == model and r.size_target == target:
            return bool(r.both_ok)
    return False


def filter_for_exp2(ctx_results: List[ContextResult]) -> List[str]:
    """Models that pass the 32k context filter (pass or unknown)."""
    return [r.model for r in ctx_results if r.status in ("pass", "unknown")]


def filter_for_exp3(
    ctx_results: List[ContextResult], probe_results: List[ProbeResult]
) -> List[str]:
    """Models eligible for experiment 3:
    - reported ctx >= 32k, OR
    - unknown ctx but passed the 64k probe
    """
    eligible = []
    for r in ctx_results:
        if r.status == "pass":
            eligible.append(r.model)
        elif r.status == "unknown" and _probe_passed_64k(r.model, probe_results):
            eligible.append(r.model)
    return eligible


# =============================================================================
# Terminal summary
# =============================================================================


def print_summary(
    ctx_results: List[ContextResult],
    probe_results: List[ProbeResult],
    author_results: List[AuthorResult],
) -> None:
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")

    # --- Context filter ---
    passed = [r for r in ctx_results if r.status == "pass"]
    filtered = [r for r in ctx_results if r.status == "filtered"]
    unknown = [r for r in ctx_results if r.status == "unknown"]

    print(f"\n=== CONTEXT FILTER (threshold: {CONTEXT_THRESHOLD:,} tokens) ===")
    print(f"  Models with ctx >= 32k : {len(passed)}")
    print(f"  Models filtered (<32k) : {len(filtered)}")
    print(f"  Models unknown         : {len(unknown)}")

    if filtered:
        print("  Filtered models:")
        for r in filtered:
            print(f"    - {r.model}  ({r.ctx_int:,})")

    # --- 64k probe results by family ---
    target_64k = 65536
    probe_64k = {r.model: r for r in probe_results if r.size_target == target_64k}

    # Gather all models that ran experiment 2
    exp2_models = list(dict.fromkeys(r.model for r in probe_results))

    family_stats: Dict[str, Dict[str, int]] = {}
    for model in exp2_models:
        fam = classify_family(model)
        if fam not in family_stats:
            family_stats[fam] = {"total": 0, "pass_64k": 0}
        family_stats[fam]["total"] += 1
        pr = probe_64k.get(model)
        if pr and pr.both_ok:
            family_stats[fam]["pass_64k"] += 1

    print(f"\n=== 64k PROBE RESULTS BY FAMILY ===")
    for fam in sorted(family_stats):
        s = family_stats[fam]
        print(
            f"  {fam:<12} {s['pass_64k']:2}/{s['total']:2} passed 64k probe (both keys)"
        )

    # --- Top 5 author extraction ---
    eligible_models = {
        r.model for r in probe_results if r.both_ok and r.size_target == target_64k
    }
    # Also include models that weren't probed at 64k but have ctx >= 32k
    for r in ctx_results:
        if r.status == "pass" and r.model not in probe_64k:
            eligible_models.add(r.model)

    eligible_authors = [
        r for r in author_results if r.model in eligible_models and not r.error
    ]
    eligible_authors.sort(key=lambda r: r.score_pct, reverse=True)

    print(f"\n=== TOP 5 AUTHOR EXTRACTION (64k probe passers) ===")
    if not eligible_authors:
        print("  No eligible results.")
    else:
        for rank, r in enumerate(eligible_authors[:5], 1):
            lat = f"{r.latency:.1f}s" if r.latency else "N/A"
            print(f"  {rank}. {r.model:<45} {r.score_pct:5.1f}%   ({lat})")

    print(f"\n{'='*60}\n")


# =============================================================================
# Main
# =============================================================================


def main() -> int:
    cfg, csv_path = parse_args()

    # Load study text
    study_path = Path(cfg.study_file)
    if not study_path.exists():
        print(f"ERROR: study file not found: {cfg.study_file}", file=sys.stderr)
        return 1
    study_text = study_path.read_text(encoding="utf-8")
    print(f"Study file loaded: {study_path.name} ({len(study_text):,} chars)")

    # Load model list
    models = load_models(csv_path)
    print(f"Models to evaluate: {len(models)}")
    print(f"Base URL: {cfg.base_url}")
    print(f"Delay between generate calls: {cfg.delay}s")
    print(f"Request timeout: {cfg.timeout}s")
    print(f"Output: {cfg.output}")

    # -------------------------------------------------------------------------
    # Load any previously saved results (resume support)
    # -------------------------------------------------------------------------
    all_exp1, all_exp2, all_exp3 = load_existing_results(cfg.output)
    if any([all_exp1, all_exp2, all_exp3]):
        print(
            f"\nResuming from {cfg.output}: "
            f"{len(all_exp1)} exp1 rows, {len(all_exp2)} exp2 rows, "
            f"{len(all_exp3)} exp3 rows already saved."
        )
    else:
        print(f"\nStarting fresh run → {cfg.output}")

    # -------------------------------------------------------------------------
    # Experiment 1: context window
    # -------------------------------------------------------------------------
    run_experiment_1(models, cfg, all_exp1, all_exp2, all_exp3)
    ctx_by_model = {r.model: r for r in all_exp1}

    # -------------------------------------------------------------------------
    # Experiment 2: needle probe
    # -------------------------------------------------------------------------
    exp2_models = filter_for_exp2(all_exp1)
    print(f"\nModels entering Experiment 2 (probe): {len(exp2_models)}")
    if exp2_models:
        run_experiment_2(exp2_models, ctx_by_model, cfg, all_exp1, all_exp2, all_exp3)
    else:
        print("  Skipping — no models passed context filter.")

    # -------------------------------------------------------------------------
    # Experiment 3: author extraction
    # -------------------------------------------------------------------------
    exp3_models = filter_for_exp3(all_exp1, all_exp2)
    print(f"\nModels entering Experiment 3 (extraction): {len(exp3_models)}")
    if exp3_models:
        run_experiment_3(exp3_models, study_text, cfg, all_exp1, all_exp2, all_exp3)
    else:
        print("  Skipping — no models passed context + probe filter.")

    # -------------------------------------------------------------------------
    # Final save + terminal summary
    # -------------------------------------------------------------------------
    write_excel(cfg.output, all_exp1, all_exp2, all_exp3)
    print_summary(all_exp1, all_exp2, all_exp3)

    return 0


if __name__ == "__main__":
    sys.exit(main())
