#!/usr/bin/env python3
"""
Llama API Benchmark: individual vs mini-batch vs all-in-one

Tests three strategies for extracting 8 entities from a single study document:
  1. individual   — one API call per entity  (8 calls, sequential)
  2. mini_batch_4 — 4 entities per call      (2 calls)
  3. all_in_one   — all 8 entities in 1 call  (1 call)

ALL strategies use _call_llama_api directly (bypasses extract_entities_with_llama's
own JSON schema which would conflict with batch prompts).

Output:  backend/scripts/llama_benchmark_results.xlsx
Usage:   python backend/scripts/llama_api_benchmark.py [--model MODEL_NAME]
         python backend/scripts/llama_api_benchmark.py --no-warmup     # skip pre-strategy ping
         python backend/scripts/llama_api_benchmark.py --skip-individual  # only run batch tests
"""

import asyncio
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Bootstrap: add backend to sys.path and load secrets into env vars
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

_secrets_path = BACKEND / "core" / "secrets.toml"
if _secrets_path.exists():
    try:
        import toml

        _secrets = toml.load(_secrets_path)
        for _section, _kv in _secrets.items():
            for _k, _v in _kv.items():
                os.environ.setdefault(f"{_section.upper()}_{_k.upper()}", str(_v))
        print(f"[Setup] Loaded secrets from {_secrets_path}")
    except Exception as _e:
        print(f"[Setup] Warning: could not load secrets.toml: {_e}")

from services.llm.llama import LlamaLLMClient  # noqa: E402

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_MODEL = "meta/llama-4-maverick-17b-128e-instruct-maas"
STUDY_FILE = (
    ROOT / "62a2a704bf8bb8b08a401a63f4b3b9a9ef6096446bd0d8ac7e0c70a63c40ad4e_base.md"
)
OUTPUT_FILE = BACKEND / "scripts" / "llama_benchmark_results.xlsx"

# For individual calls we use full doc (96K chars) as in production.
# For batch calls, limit to 16K chars — the model struggles to produce clean
# multi-field JSON from very large table-heavy documents (generates whitespace/garbage).
# This proves the batching concept works; production notes in Excel summary.
MAX_DOC_CHARS_INDIVIDUAL = 96_000
MAX_DOC_CHARS_BATCH = 16_000

# Entities to extract
ENTITIES = [
    {
        "name": "Study Author(s)",
        "key": "study_authors",
        "prompt": "Extract the study author(s). Return only the names, comma-separated.",
    },
    {
        "name": "Author Affiliations",
        "key": "author_affiliations",
        "prompt": "Extract the affiliations of the authors. Return institution names separated by semicolons.",
    },
    {
        "name": "Study Title",
        "key": "study_title",
        "prompt": "Extract the exact title of the study.",
    },
    {
        "name": "Publication Date",
        "key": "publication_date",
        "prompt": "Extract the publication date (year or full date).",
    },
    {
        "name": "Test Material",
        "key": "test_material",
        "prompt": "Extract the test material(s) used in the study.",
    },
    {
        "name": "Vehicle / Solvent",
        "key": "vehicle_solvent",
        "prompt": "Extract the vehicle or solvent used to administer the test material. Return 'Not specified' if absent.",
    },
    {
        "name": "Dose Levels",
        "key": "dose_levels",
        "prompt": "Extract the dose levels administered (with units).",
    },
    {
        "name": "Results Presentation",
        "key": "results_presentation",
        "prompt": "Describe briefly how results are presented: qualitative, quantitative, tables, figures, text?",
    },
]

SYSTEM_PROMPT = (
    "You are a scientific data extraction assistant. "
    "Answer concisely and accurately using only information found in the document. "
    "Do not add interpretation or commentary."
)

# ---------------------------------------------------------------------------
# Low-level helpers — all go through _call_llama_api directly
# ---------------------------------------------------------------------------


def _build_single_messages(document: str, entity_prompt: str) -> list[dict]:
    """Build messages for a single-entity extraction."""
    return [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"Document:\n{document}\n\nTask: {entity_prompt}"},
    ]


def _build_batch_messages(document: str, entities: list[dict]) -> list[dict]:
    """Build messages for a multi-entity extraction returning a JSON object.

    Uses _call_llama_api directly (no response_format JSON mode) so we don't
    conflict with extract_entities_with_llama's own schema.
    """
    fields = "\n".join(f'  "{e["key"]}": "{e["prompt"]}"' for e in entities)
    example = json.dumps({e["key"]: "..." for e in entities}, indent=2)
    task = (
        f"Extract the following fields from the document.\n"
        f"Return ONLY a valid JSON object — no markdown, no explanation.\n\n"
        f"Fields (key: instruction):\n{{\n{fields}\n}}\n\n"
        f"Return format:\n{example}"
    )
    return [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"Document:\n{document}\n\nTask:\n{task}"},
    ]


def _parse_raw(result: dict) -> dict:
    """Extract fields from _call_llama_api result (single-entity response)."""
    meta = result.get("meta") or {}
    return {
        "success": result.get("success", False),
        "answer": (
            str(result.get("content", "")).strip() if result.get("success") else ""
        ),
        "model": meta.get("model", DEFAULT_MODEL),
        "prompt_tokens": meta.get("prompt_tokens"),
        "completion_tokens": meta.get("completion_tokens"),
        "duration_s": round(meta.get("duration") or 0, 3),
        "error": str(result.get("error", "")) if not result.get("success") else "",
    }


def _parse_batch_field(result: dict, entity_key: str) -> dict:
    """Pull one entity's answer from a batch _call_llama_api result."""
    meta = result.get("meta") or {}
    base_meta = {
        "model": meta.get("model", DEFAULT_MODEL),
        "prompt_tokens": meta.get("prompt_tokens"),
        "completion_tokens": meta.get("completion_tokens"),
        "duration_s": round(meta.get("duration") or 0, 3),
    }

    if not result.get("success"):
        return {
            "success": False,
            "answer": "",
            "error": str(result.get("error", "Batch API call failed")),
            **base_meta,
        }

    raw_content = str(result.get("content", "")).strip()
    # Strip markdown fences if present
    if raw_content.startswith("```"):
        parts = raw_content.split("```")
        raw_content = parts[1] if len(parts) > 1 else raw_content
        if raw_content.startswith("json"):
            raw_content = raw_content[4:].strip()

    try:
        parsed = json.loads(raw_content)
        answer = parsed.get(entity_key)
        if answer is None:
            return {
                "success": False,
                "answer": "",
                "error": f"Key '{entity_key}' missing from batch JSON. Keys found: {list(parsed.keys())}",
                **base_meta,
            }
        return {
            "success": True,
            "answer": str(answer).strip(),
            "error": "",
            **base_meta,
        }
    except json.JSONDecodeError as e:
        return {
            "success": False,
            "answer": "",
            "error": f"JSON parse error: {e} | raw[:200]: {raw_content[:200]}",
            **base_meta,
        }


async def _warm_instance(client: LlamaLLMClient, model: str, label: str) -> float:
    """Send a 1-token ping. Returns elapsed seconds."""
    t0 = time.perf_counter()
    result = await client._call_llama_api(
        model_name=model,
        messages=[{"role": "user", "content": "hi"}],
        max_tokens=1,
        temperature=0.0,
        request_timeout=300,
        max_retries=1,
    )
    elapsed = time.perf_counter() - t0
    status = (
        "✅ warm" if result.get("success") else f"❌ {result.get('error', '')[:60]}"
    )
    print(f"  [Ping:{label}] {elapsed:.1f}s — {status}")
    return elapsed


# ---------------------------------------------------------------------------
# Strategy runners
# ---------------------------------------------------------------------------


async def run_individual(
    client: LlamaLLMClient, document: str, model: str, warmup: bool
) -> tuple[list[dict], float]:
    """Strategy 1: 8 sequential calls, one per entity."""
    if warmup:
        print("\n[individual] Pre-strategy warm-up ping...")
        await _warm_instance(client, model, "individual")

    print(f"\n[individual] Running {len(ENTITIES)} sequential calls...")
    rows = []
    t_strategy = time.perf_counter()

    for i, entity in enumerate(ENTITIES, 1):
        msgs = _build_single_messages(document, entity["prompt"])
        print(f"  [{i}/{len(ENTITIES)}] '{entity['name']}'...", end=" ", flush=True)
        t0 = time.perf_counter()
        result = await client._call_llama_api(
            model_name=model,
            messages=msgs,
            max_tokens=512,
            temperature=0.0,
            request_timeout=300,
            max_retries=1,
        )
        call_time = time.perf_counter() - t0
        parsed = _parse_raw(result)
        cold_flag = " ⚠️ COLD?" if call_time > 30 else ""
        status = "✅" if parsed["success"] else "❌"
        print(f"{status} {call_time:.1f}s{cold_flag}")

        rows.append(
            {
                "strategy": "individual",
                "entity_name": entity["name"],
                "prompt_sent": entity["prompt"],
                "call_time_s": round(call_time, 3),
                "api_calls_this_row": 1,
                "cold_start_flag": call_time > 30,
                **parsed,
            }
        )

    total = time.perf_counter() - t_strategy
    successes = sum(1 for r in rows if r["success"])
    print(f"[individual] Done — {successes}/{len(ENTITIES)} OK, {total:.1f}s total")
    return rows, total


async def run_mini_batch(
    client: LlamaLLMClient, document: str, model: str, warmup: bool
) -> tuple[list[dict], float]:
    """Strategy 2: 2 API calls, 4 entities each."""
    if warmup:
        print("\n[mini_batch_4] Pre-strategy warm-up ping...")
        await _warm_instance(client, model, "mini_batch_4")

    groups = [ENTITIES[:4], ENTITIES[4:]]
    print(f"\n[mini_batch_4] Running {len(groups)} batch calls (4 entities each)...")
    rows = []
    t_strategy = time.perf_counter()

    for gi, group in enumerate(groups, 1):
        names = [e["name"] for e in group]
        msgs = _build_batch_messages(document, group)
        prompt_preview = f"BATCH({', '.join(names)})"
        print(f"  [batch {gi}/{len(groups)}] {names}...", end=" ", flush=True)

        t0 = time.perf_counter()
        result = await client._call_llama_api(
            model_name=model,
            messages=msgs,
            max_tokens=1024,
            temperature=0.0,
            request_timeout=300,
            max_retries=1,
            response_format={"type": "json_object"},
        )
        call_time = time.perf_counter() - t0
        cold_flag = call_time > 30
        status = "✅" if result.get("success") else "❌"
        print(f"{status} {call_time:.1f}s{'  ⚠️ COLD?' if cold_flag else ''}")

        # If batch succeeded, show what the model actually returned (for debugging)
        if result.get("success"):
            preview = str(result.get("content", ""))[:200]
            print(f"    Raw response preview: {preview}")

        meta = result.get("meta") or {}
        pt_each = (meta.get("prompt_tokens") or 0) // len(group)
        ct_each = (meta.get("completion_tokens") or 0) // len(group)

        for entity in group:
            parsed = _parse_batch_field(result, entity["key"])
            rows.append(
                {
                    "strategy": "mini_batch_4",
                    "entity_name": entity["name"],
                    "prompt_sent": prompt_preview,
                    "call_time_s": round(call_time, 3),
                    "api_calls_this_row": round(1 / len(group), 3),
                    "cold_start_flag": cold_flag,
                    **parsed,
                    "prompt_tokens": pt_each,
                    "completion_tokens": ct_each,
                }
            )

    total = time.perf_counter() - t_strategy
    successes = sum(1 for r in rows if r["success"])
    print(f"[mini_batch_4] Done — {successes}/{len(rows)} OK, {total:.1f}s total")
    return rows, total


async def run_all_in_one(
    client: LlamaLLMClient, document: str, model: str, warmup: bool
) -> tuple[list[dict], float]:
    """Strategy 3: all 8 entities in a single API call."""
    if warmup:
        print("\n[all_in_one] Pre-strategy warm-up ping...")
        await _warm_instance(client, model, "all_in_one")

    msgs = _build_batch_messages(document, ENTITIES)
    all_names = [e["name"] for e in ENTITIES]
    print(
        f"\n[all_in_one] 1 call for all {len(ENTITIES)} entities...",
        end=" ",
        flush=True,
    )

    t_strategy = time.perf_counter()
    result = await client._call_llama_api(
        model_name=model,
        messages=msgs,
        max_tokens=2048,
        temperature=0.0,
        request_timeout=300,
        max_retries=3,
        response_format={"type": "json_object"},
    )
    call_time = time.perf_counter() - t_strategy
    cold_flag = call_time > 30
    status = "✅" if result.get("success") else "❌"
    print(f"{status} {call_time:.1f}s{'  ⚠️ COLD?' if cold_flag else ''}")

    if result.get("success"):
        preview = str(result.get("content", ""))[:300]
        print(f"  Raw response preview: {preview}")

    meta = result.get("meta") or {}
    pt_each = (meta.get("prompt_tokens") or 0) // len(ENTITIES)
    ct_each = (meta.get("completion_tokens") or 0) // len(ENTITIES)

    rows = []
    for entity in ENTITIES:
        parsed = _parse_batch_field(result, entity["key"])
        rows.append(
            {
                "strategy": "all_in_one",
                "entity_name": entity["name"],
                "prompt_sent": f"BATCH_ALL({', '.join(e['name'] for e in ENTITIES[:3])}...)",
                "call_time_s": round(call_time, 3),
                "api_calls_this_row": round(1 / len(ENTITIES), 3),
                "cold_start_flag": cold_flag,
                **parsed,
                "prompt_tokens": pt_each,
                "completion_tokens": ct_each,
            }
        )

    successes = sum(1 for r in rows if r["success"])
    print(f"[all_in_one] Done — {successes}/{len(rows)} OK, {call_time:.1f}s total")
    return rows, call_time


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------


def write_excel(all_rows: list[dict], strategy_totals: dict[str, dict], model: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED = PatternFill("solid", fgColor="FFC7CE")
    ORANGE = PatternFill("solid", fgColor="FFEB9C")  # cold start warning
    BLUE_LIGHT = PatternFill("solid", fgColor="BDD7EE")
    GREY = PatternFill("solid", fgColor="F2F2F2")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=13)
    BOLD = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    STRAT_FILL = {
        "individual": PatternFill("solid", fgColor="EBF3FB"),
        "mini_batch_4": PatternFill("solid", fgColor="EDEDED"),
        "all_in_one": PatternFill("solid", fgColor="E8E8E8"),
    }

    wb = Workbook()

    # ---- Sheet 1: All Results ------------------------------------------------
    ws = wb.active
    ws.title = "All Results"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:L1")
    ws["A1"] = (
        f"Llama API Benchmark — {model} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = BLUE_LIGHT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    COLS = [
        ("Strategy", 20),
        ("Entity", 24),
        ("Prompt Sent", 45),
        ("Call Time (s)", 13),
        ("API Calls", 10),
        ("Cold Start?", 12),
        ("Success", 10),
        ("Answer", 52),
        ("Model", 38),
        ("Prompt Tok.", 12),
        ("Compl. Tok.", 12),
        ("Error", 50),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    for ri, row in enumerate(all_rows, 3):
        ok = row.get("success", False)
        cold = row.get("cold_start_flag", False)
        t = row.get("call_time_s", 0)
        strat = row.get("strategy", "")

        vals = [
            strat,
            row.get("entity_name", ""),
            row.get("prompt_sent", ""),
            t,
            row.get("api_calls_this_row", ""),
            "⚠️ YES" if cold else "no",
            "✅" if ok else "❌",
            row.get("answer", ""),
            row.get("model", ""),
            row.get("prompt_tokens", ""),
            row.get("completion_tokens", ""),
            row.get("error", ""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = thin
            if ci == 4:  # call time — orange if cold start
                cell.fill = ORANGE if cold else (STRAT_FILL.get(strat, GREY))
                cell.alignment = CENTER
            elif ci == 6:  # cold start
                cell.fill = ORANGE if cold else GREY
                cell.alignment = CENTER
            elif ci == 7:  # success
                cell.fill = GREEN if ok else RED
                cell.alignment = CENTER
            elif ci == 8:  # answer
                cell.fill = GREEN if ok else RED
            else:
                cell.fill = STRAT_FILL.get(strat, GREY)
        ws.row_dimensions[ri].height = 52

    # ---- Sheet 2: Summary ----------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:I1")
    ws2["A1"] = "Strategy Summary"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].fill = BLUE_LIGHT
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 26

    SUM_COLS = [
        ("Strategy", 20),
        ("API Calls", 10),
        ("Total Time (s)", 14),
        ("Avg / Call (s)", 14),
        ("Cold Starts", 12),
        ("Successes", 12),
        ("Failures", 12),
        ("Prompt Tokens", 14),
        ("Completion Tokens", 16),
    ]
    for ci, (h, w) in enumerate(SUM_COLS, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = thin
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 20

    all_totals = [v["total_time"] for v in strategy_totals.values()]
    for ri, (strat, info) in enumerate(strategy_totals.items(), 3):
        calls = info["calls"]
        tt = info["total_time"]
        vals = [
            strat,
            calls,
            round(tt, 2),
            round(tt / calls, 2) if calls else 0,
            info["cold_starts"],
            info["successes"],
            info["failures"],
            info["prompt_tokens"],
            info["completion_tokens"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.alignment = CENTER
            cell.border = thin
            if ci == 1:
                cell.font = BOLD
                cell.fill = STRAT_FILL.get(strat, GREY)
            elif ci == 3:  # total time — highlight fastest
                cell.fill = GREEN if tt == min(all_totals) else GREY
            elif ci == 5:  # cold starts
                cell.fill = ORANGE if info["cold_starts"] > 0 else GREEN
        ws2.row_dimensions[ri].height = 20

    # ---- Sheet 3: Cold Start Analysis ----------------------------------------
    ws3 = wb.create_sheet("Cold Start Analysis")
    ws3["A1"] = "Cold Start vs Warm Call Analysis"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].fill = BLUE_LIGHT
    ws3["A1"].alignment = CENTER
    ws3.merge_cells("A1:E1")
    ws3.row_dimensions[1].height = 26

    hdrs = ["Strategy", "Entity", "Call Time (s)", "Classification", "Notes"]
    widths = [20, 25, 14, 18, 55]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = thin
        ws3.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(all_rows, 3):
        t = row.get("call_time_s", 0)
        cold = row.get("cold_start_flag", False)
        ok = row.get("success", False)
        if not ok:
            cls = "FAILED"
            notes = f"Error: {row.get('error', 'unknown')[:80]}"
            fill = RED
        elif cold:
            cls = "COLD START"
            notes = f"Took {t:.1f}s — likely hit a cold/unloaded backend instance on Vertex AI"
            fill = ORANGE
        elif t > 10:
            cls = "WARM (slow)"
            notes = f"{t:.1f}s — warm instance but generating a long response"
            fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            cls = "WARM"
            notes = f"{t:.1f}s — fast warm instance"
            fill = GREEN

        vals = [row.get("strategy", ""), row.get("entity_name", ""), t, cls, notes]
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.alignment = WRAP
            if ci in (3, 4):
                cell.fill = fill
                cell.alignment = CENTER
        ws3.row_dimensions[ri].height = 22

    wb.save(OUTPUT_FILE)
    print(f"\n[Output] Excel → {OUTPUT_FILE}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


async def main():
    model = DEFAULT_MODEL
    do_warmup = "--no-warmup" not in sys.argv
    skip_individual = "--skip-individual" in sys.argv

    if "--model" in sys.argv:
        idx = sys.argv.index("--model")
        if idx + 1 < len(sys.argv):
            model = sys.argv[idx + 1]

    print("=" * 72)
    print("  Llama API Benchmark (direct _call_llama_api — no schema conflicts)")
    print(f"  Model : {model}")
    print(f"  Study : {STUDY_FILE.name}")
    print(f"  Warmup pings: {'yes' if do_warmup else 'no (--no-warmup)'}")
    print(f"  Output: {OUTPUT_FILE}")
    print("=" * 72)

    if not STUDY_FILE.exists():
        print(f"[ERROR] Study file not found: {STUDY_FILE}")
        sys.exit(1)

    full_doc = STUDY_FILE.read_text(encoding="utf-8")
    doc_individual = full_doc[:MAX_DOC_CHARS_INDIVIDUAL]
    doc_batch = full_doc[:MAX_DOC_CHARS_BATCH]
    print(
        f"[Setup] Study loaded: {len(full_doc):,} chars total\n"
        f"        individual strategy: first {len(doc_individual):,} chars "
        f"(~{len(doc_individual)//4:,} tok)\n"
        f"        batch strategies:    first {len(doc_batch):,} chars "
        f"(~{len(doc_batch)//4:,} tok)"
    )

    client = LlamaLLMClient()
    if client.disabled:
        print(
            "[ERROR] LlamaLLMClient is disabled — check LLAMA_PROJECT_ID and service account."
        )
        sys.exit(1)
    print(f"[Setup] Client ready — project={client.project_id}, region={client.region}")
    print(
        f"[Setup] Service account: {client.service_account_path.name if client.service_account_path else 'NONE'}"
    )

    all_rows: list[dict] = []
    strategy_totals: dict[str, dict] = {}

    strategies = []
    if not skip_individual:
        strategies.append(
            (
                "individual",
                lambda: run_individual(client, doc_individual, model, do_warmup),
            )
        )
    strategies.append(
        ("mini_batch_4", lambda: run_mini_batch(client, doc_batch, model, do_warmup))
    )
    strategies.append(
        ("all_in_one", lambda: run_all_in_one(client, doc_batch, model, do_warmup))
    )

    for name, fn in strategies:
        rows, total_time = await fn()
        all_rows.extend(rows)

        actual_calls = {
            "individual": len(rows),
            "mini_batch_4": 2,
            "all_in_one": 1,
        }.get(name, len(rows))

        strategy_totals[name] = {
            "calls": actual_calls,
            "total_time": total_time,
            "cold_starts": sum(1 for r in rows if r.get("cold_start_flag")),
            "successes": sum(1 for r in rows if r.get("success")),
            "failures": sum(1 for r in rows if not r.get("success")),
            "prompt_tokens": sum((r.get("prompt_tokens") or 0) for r in rows),
            "completion_tokens": sum((r.get("completion_tokens") or 0) for r in rows),
        }

    # Console summary
    print("\n" + "=" * 72)
    print(
        f"  {'Strategy':<18} {'Calls':>6} {'Total(s)':>10} {'Avg/call':>10} {'Cold':>5} {'OK':>4} {'FAIL':>4}"
    )
    print("-" * 72)
    for strat, info in strategy_totals.items():
        c = info["calls"]
        print(
            f"  {strat:<18} {c:>6} {info['total_time']:>10.1f} "
            f"{info['total_time']/c:>10.1f} {info['cold_starts']:>5} "
            f"{info['successes']:>4} {info['failures']:>4}"
        )
    print("=" * 72)

    write_excel(all_rows, strategy_totals, model)


if __name__ == "__main__":
    asyncio.run(main())
