#!/usr/bin/env python3
"""
Multi-Model LLM Evaluation Script

This script evaluates multiple LLMs hosted on a MacBook-based inference cluster.
It sends per-entity extraction prompts to each model, measures latency, evaluates
responses against ground truth, and outputs results to an Excel file.

Usage:
    python backend/scripts/evaluate_macbook_models.py

Environment Variables:
    MACBOOK_LLM_BASE_URL: Base URL for the MacBook LLM API (default: http://macbook1.sciencegpt.ca)
    DELAY_BETWEEN_REQUESTS: Delay in seconds between model requests (default: 60, max wait)
    MAX_TOKENS: Maximum tokens in response (default: 4096)
    TEMPERATURE: Temperature for generation (default: 0.0)
    REQUEST_TIMEOUT: Timeout per request in seconds (default: 600)
"""

import os
import sys
import time
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
import requests
import toml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =============================================================================
# Configuration
# =============================================================================

@dataclass
class Config:
    """Configuration settings for the evaluation script."""
    
    # MacBook API settings
    base_url: str = "http://macbook1.sciencegpt.ca"
    api_endpoint: str = "/api/generate"
    
    # Request settings
    delay_between_requests: int = 60  # Maximum delay in seconds
    max_tokens: int = 4096
    temperature: float = 0.0
    request_timeout: int = 600  # seconds
    
    # File paths (relative to project root)
    project_root: Path = Path(__file__).resolve().parents[2]
    model_list_file: Path = field(default_factory=lambda: Path("macbookmodelnames.csv"))
    prompt_file: Path = field(default_factory=lambda: Path("prompt.md"))
    test_document_file: Path = field(default_factory=lambda: Path("64596011f75ffd2916b1ce50131f3d7cb36c10141e914e435fd5dc0e007b2b52_base.md"))
    output_file: Path = field(default_factory=lambda: Path("model_evaluation_results.xlsx"))
    
    @classmethod
    def from_env(cls) -> 'Config':
        """Create configuration from environment variables."""
        base_url = os.environ.get("MACBOOK_LLM_BASE_URL")
        if not base_url:
            base_url = cls._load_base_url_from_secrets()
        
        return cls(
            base_url=base_url or "http://macbook1.sciencegpt.ca",
            delay_between_requests=int(os.environ.get("DELAY_BETWEEN_REQUESTS", 60)),
            max_tokens=int(os.environ.get("MAX_TOKENS", 4096)),
            temperature=float(os.environ.get("TEMPERATURE", 0.0)),
            request_timeout=int(os.environ.get("REQUEST_TIMEOUT", 600)),
        )
    
    @staticmethod
    def _load_base_url_from_secrets() -> Optional[str]:
        """Load base URL from secrets.toml."""
        try:
            secrets_path = Path(__file__).resolve().parents[2] / "core" / "secrets.toml"
            if secrets_path.exists():
                data = toml.load(secrets_path)
                macbook_section = data.get("Macbook") or {}
                return macbook_section.get("macbook_llm_base_url")
        except Exception as e:
            print(f"Warning: Could not load secrets.toml: {e}")
        return None


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class TestResult:
    """Result of testing a single model."""
    model_name: str
    entity_responses: Dict[str, str]  # entity_name -> extracted_response
    ground_truth: Dict[str, str]  # entity_name -> expected_answer
    score: float
    total_latency_seconds: float
    error: Optional[str] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for Excel export."""
        return {
            "Model Name": self.model_name,
            "Entity Outputs": "\n\n".join([f"{k}: {v}" for k, v in self.entity_responses.items()]),
            "Ground Truth": "\n\n".join([f"{k}: {v}" for k, v in self.ground_truth.items()]),
            "Score": f"{self.score:.2%}" if self.error is None else "N/A",
            "Latency (seconds)": f"{self.total_latency_seconds:.2f}" if self.error is None else "N/A",
            "Error": self.error or ""
        }


@dataclass
class EntityPrompt:
    """Represents a single entity extraction prompt."""
    entity_name: str
    prompt_with_examples: str  # Includes few-shot examples


# =============================================================================
# Model List Parser
# =============================================================================

class ModelListParser:
    """Parse model names from the CSV file."""
    
    EXPECTED_MODELS = [
        # 3B - 4B
        "llama3.2:3b-instruct-fp16",
        "llama3.2:3b-instruct-q4_K_M",
        "MedAIBase/MedGemma1.5:4b",
        "phi4-mini:3.8b",
        "phi3.5:3.8b",
        "nemotron-mini:4b-instruct-q4_K_M",
        "nemotron-mini:4b-instruct-fp16",
        "nemotron-mini:4b-instruct-q8_0",
        
        # 7B - 8B
        "llama3.1:8b",
        "llama3.1:8b-instruct-q4_K_M",
        "dolphin-llama3:8b",
        "openbiollm-llama-3:8b-q8_0",
        "openbiollm-llama-3:8b-q6_k",
        "Mistral-7B-Instruct-v0.3-Q4_K_M:latest",
        "qwen3:8b-q4_K_M",
        "rnj-1:8b",
        
        # 12B - 14B
        "ministral-3:14b-instruct-2512-q4_K_M",
        "mistral-nemo:12b",
        "mistral-nemo:12b-instruct-2407-q3_K_M",
        "gemma3:12b",
        "qwen3:14b-q4_K_M",
        "phi4-reasoning:14b",
        
        # 20B - 24B
        "mistral-small3.1:24b",
        "gpt-o3:20b",
        
        # 27B - 32B
        "gemma3:27b-it-qat",
        "gemma3:27b-it-q4_K_M",
        "gemma2:27b",
        "qwen3:30b-a3b-q4_K_M",
        "qwen2.5:32b",
        "nemotron-3-nano:30b",
        "olmo-2:32b-think-q4_K_M",
        
        # 70B+
        "llama3.3:70b-instruct-q2_K",
        "llama3.1:70b-instruct-q2_k",
        "llama3.1:70b",
        "openbiollm-llama-3:70b_q4_k_m",
        "mistral-large:123b",
        "nemotron:70b-instruct-q3_K_M",
        "nemotron:70b-instruct-q2_K",
    ]
    
    @classmethod
    def parse_model_list(cls, csv_content: str) -> List[str]:
        """Parse model names from CSV content."""
        valid_models = []
        
        for model in cls.EXPECTED_MODELS:
            if model.lower().replace('-', '').replace(':', '') in csv_content.lower().replace('-', '').replace(':', ''):
                valid_models.append(model)
        
        if len(valid_models) < 10:
            return cls.EXPECTED_MODELS
        
        return sorted(valid_models)


# =============================================================================
# Prompt and Document Loader
# =============================================================================

class PromptLoader:
    """Load prompt template and test document."""
    
    def __init__(self, config: Config):
        self.config = config
    
    def load(self) -> Tuple[List[EntityPrompt], Dict[str, str]]:
        """Load prompt and document, return entity prompts and ground truth."""
        # Load prompt file
        prompt_path = self.config.project_root / self.config.prompt_file
        with open(prompt_path, 'r', encoding='utf-8') as f:
            prompt_content = f.read()
        
        # Parse entity prompts and ground truth
        entity_prompts, ground_truth = self._parse_prompt_file(prompt_content)
        
        # Load test document
        doc_path = self.config.project_root / self.config.test_document_file
        with open(doc_path, 'r', encoding='utf-8') as f:
            document_content = f.read()
        
        # Build full prompts with document
        for ep in entity_prompts:
            ep.prompt_with_examples = self._build_full_prompt(ep.prompt_with_examples, document_content)
        
        return entity_prompts, ground_truth
    
    def _parse_prompt_file(self, content: str) -> Tuple[List[EntityPrompt], Dict[str, str]]:
        """Parse prompt.md to extract entity prompts and ground truth."""
        # Split by "Ground Truth:" section
        parts = content.split("Ground Truth:")
        
        prompts_section = parts[0].strip()
        ground_truth_section = parts[1].strip() if len(parts) > 1 else ""
        
        # Parse individual entity prompts (separated by dashes)
        entity_prompts = []
        prompt_blocks = re.split(r'-{5,}', prompts_section)
        
        for block in prompt_blocks:
            block = block.strip()
            if not block or block.startswith("Few-shot"):
                continue
            
            # Extract entity name from "Extract the..." line
            lines = block.split('\n')
            entity_line = ""
            prompt_lines = []
            
            for line in lines:
                line = line.strip()
                if line.startswith("Extract "):
                    entity_line = line
                elif line and not line.startswith("Input:") and not line.startswith("Output:"):
                    prompt_lines.append(line)
            
            if entity_line:
                # Create entity name from prompt
                entity_name = self._extract_entity_name(entity_line)
                full_prompt = "\n".join(prompt_lines)
                
                if entity_name and full_prompt:
                    entity_prompts.append(EntityPrompt(
                        entity_name=entity_name,
                        prompt_with_examples=full_prompt
                    ))
        
        # Parse ground truth
        ground_truth = self._parse_ground_truth(ground_truth_section)
        
        return entity_prompts, ground_truth
    
    def _extract_entity_name(self, prompt_line: str) -> str:
        """Extract entity name from prompt line."""
        # Remove "Extract the" prefix and trailing period
        name = prompt_line.replace("Extract the", "").replace("Extract", "").strip()
        name = re.sub(r'\.$', '', name).strip()
        
        # Normalize to lowercase with underscores
        name = name.lower().replace(" ", "_").replace("-", "_")
        
        return name
    
    def _parse_ground_truth(self, gt_section: str) -> Dict[str, str]:
        """Parse ground truth section into key-value pairs."""
        ground_truth = {}
        
        # Split by dashes
        gt_blocks = re.split(r'-{5,}', gt_section)
        
        # Entity names that match prompt parsing
        entity_names = [
            "study_author(s)",
            "author_affiliations",
            "study_title",
            "publication_date",
            "test_material",
            "vehicle_or_solvent_used",
            "dose_levels",
            "how_the_results_are_presented"
        ]
        
        for i, block in enumerate(gt_blocks):
            block = block.strip()
            if not block:
                continue
            
            if i < len(entity_names):
                ground_truth[entity_names[i]] = block
        
        return ground_truth
    
    def _build_full_prompt(self, entity_prompt: str, document: str) -> str:
        """Build the full prompt with document content."""
        return f"""<markdown study>
{document}
</markdown study>

{entity_prompt}

Output only the extracted information, nothing else."""


# =============================================================================
# MacBook Client
# =============================================================================

class MacbookClient:
    """Client for interacting with MacBook LLM API."""
    
    def __init__(self, config: Config):
        self.config = config
        self.base_url = config.base_url.rstrip('/')
        self.endpoint = config.api_endpoint
        self.session = requests.Session()
    
    def generate(self, model: str, prompt: str) -> Tuple[str, float, Optional[str]]:
        """Generate response from model. Returns (response, latency, error)."""
        url = f"{self.base_url}{self.endpoint}"
        
        payload = {
            "model": model,
            "prompt": prompt,
            "stream": False,
            "temperature": self.config.temperature,
            "max_tokens": self.config.max_tokens,
        }
        
        start_time = time.time()
        
        try:
            response = self.session.post(
                url,
                json=payload,
                timeout=self.config.request_timeout
            )
            
            latency = time.time() - start_time
            
            if response.ok:
                data = response.json()
                content = data.get("response") or data.get("content") or data.get("text", "")
                return content, latency, None
            else:
                error_msg = f"HTTP {response.status_code}: {response.text[:200]}"
                return "", latency, error_msg
                
        except requests.exceptions.Timeout:
            latency = time.time() - start_time
            return "", latency, f"Request timeout after {self.config.request_timeout}s"
        except requests.exceptions.RequestException as e:
            latency = time.time() - start_time
            return "", latency, str(e)
    
    def close(self):
        """Close the session."""
        self.session.close()


# =============================================================================
# Response Evaluator
# =============================================================================

class ResponseEvaluator:
    """Evaluate model responses against ground truth."""
    
    def __init__(self, ground_truth: Dict[str, str]):
        self.ground_truth = ground_truth
        
        # Weights for 8 entities (sum to 1.0) - must match actual entity names
        self.weights = {
            "study_author(s)": 0.12,
            "author_affiliations": 0.12,
            "study_title": 0.12,
            "publication_date": 0.10,
            "test_material": 0.12,
            "vehicle_or_solvent_used": 0.12,
            "dose_levels": 0.15,
            "how_the_results_are_presented": 0.15,
        }
    
    def compute_score(self, entity_responses: Dict[str, str]) -> float:
        """Compute score based on entity responses."""
        total_score = 0.0
        
        for entity_name, response in entity_responses.items():
            weight = self.weights.get(entity_name, 0.10)
            expected = self.ground_truth.get(entity_name, "")
            
            if not expected:
                continue
            
            expected_lower = expected.lower().strip()
            response_lower = response.lower().strip()
            
            # Check for "not reported" in ground truth
            if expected_lower == "not reported" or expected_lower == "na":
                total_score += weight
                continue
            
            # Exact match
            if expected_lower in response_lower:
                total_score += weight
            else:
                # Partial match
                key_terms = self._extract_key_terms(expected)
                if key_terms:
                    matched = sum(1 for term in key_terms if term in response_lower)
                    partial = (matched / len(key_terms)) * weight
                    total_score += partial
        
        return min(total_score, 1.0)
    
    def _extract_key_terms(self, text: str) -> List[str]:
        """Extract key terms from text."""
        stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
                      'of', 'with', 'by', 'from', 'is', 'was', 'are', 'were', 'been', 'be',
                      'as', 'that', 'which', 'it', 'its', 'were', 'had', 'has', 'have'}
        
        words = re.split(r'[\s,\.\;\:\(\)\[\]\{\}]+', text.lower())
        key_terms = [w for w in words if len(w) > 2 and w not in stop_words]
        
        return key_terms


# =============================================================================
# Excel Reporter
# =============================================================================

class ExcelReporter:
    """Generate Excel report from test results."""
    
    def __init__(self, config: Config):
        self.config = config
    
    def generate(self, results: List[TestResult]):
        """Generate Excel file from test results."""
        output_path = self.config.project_root / self.config.output_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Model Evaluation Results"
        
        headers = ["Model Name", "Entity Outputs", "Ground Truth", "Score", "Latency (seconds)", "Error"]
        
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result.model_name)
            
            # Entity outputs
            outputs = "\n\n".join([f"{k}: {v[:500]}..." if len(v) > 500 else f"{k}: {v}" 
                                  for k, v in result.entity_responses.items()])
            ws.cell(row=row_idx, column=2, value=outputs)
            
            # Ground truth
            gt = "\n\n".join([f"{k}: {v[:200]}..." if len(v) > 200 else f"{k}: {v}" 
                             for k, v in result.ground_truth.items()])
            ws.cell(row=row_idx, column=3, value=gt)
            
            # Score
            ws.cell(row=row_idx, column=4, value=result.score if result.error is None else "N/A")
            
            # Latency
            ws.cell(row=row_idx, column=5, value=result.total_latency_seconds if result.error is None else "N/A")
            
            # Error
            ws.cell(row=row_idx, column=6, value=result.error or "")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 40
        
        ws.freeze_panes = 'A2'
        
        wb.save(output_path)
        print(f"\n✅ Excel report saved to: {output_path}")
        
        return output_path


# =============================================================================
# Main Test Runner
# =============================================================================

class ModelTestRunner:
    """Orchestrates the model testing process."""
    
    def __init__(self, config: Config):
        self.config = config
        self.client = MacbookClient(config)
        self.reporter = ExcelReporter(config)
    
    def run(self):
        """Run the complete evaluation process."""
        print("=" * 70)
        print("Multi-Model LLM Evaluation (Per-Entity Extraction)")
        print("=" * 70)
        
        # Load model list
        print("\n📋 Loading model list...")
        models = self._load_models()
        print(f"   Found {len(models)} models to test")
        
        # Load prompts and document
        print("\n📄 Loading prompts and document...")
        prompt_loader = PromptLoader(self.config)
        entity_prompts, ground_truth = prompt_loader.load()
        print(f"   Loaded {len(entity_prompts)} entity prompts")
        print(f"   Entities: {[ep.entity_name for ep in entity_prompts]}")
        
        # Create evaluator
        evaluator = ResponseEvaluator(ground_truth)
        
        # Calculate total calls
        total_calls = len(models) * len(entity_prompts)
        print(f"\n🚀 Starting evaluation...")
        print(f"   Total API calls: {total_calls} ({len(models)} models × {len(entity_prompts)} entities)")
        print(f"   Max delay between calls: {self.config.delay_between_requests}s")
        print("-" * 70)
        
        results = []
        
        for model_idx, model in enumerate(models, 1):
            print(f"\n[{model_idx}/{len(models)}] Testing model: {model}")
            
            entity_responses = {}
            total_latency = 0.0
            model_error = None
            
            # Per-entity extraction
            for ep in entity_prompts:
                entity_name = ep.entity_name
                
                # Log prompt details for debugging
                prompt_len = len(ep.prompt_with_examples)
                prompt_tokens_approx = prompt_len // 4  # Rough estimate
                print(f"   📤 Extracting: {entity_name}... (prompt: {prompt_len} chars, ~{prompt_tokens_approx} tokens)")
                
                # Send request
                response, latency, error = self.client.generate(model, ep.prompt_with_examples)
                
                # Log response details
                if not error:
                    response_len = len(response)
                    print(f"      📥 Response: {response_len} chars")
                
                if error:
                    print(f"      ❌ Error: {error}")
                    model_error = error
                    entity_responses[entity_name] = f"ERROR: {error}"
                else:
                    print(f"      ✅ Latency: {latency:.2f}s")
                    entity_responses[entity_name] = response.strip()
                
                total_latency += latency
                
                # Mandatory 30-second delay between calls
                # Use the max of 30 seconds and the response latency
                delay = max(30, latency)
                print(f"      ⏳ Waiting {delay:.1f}s before next request...")
                time.sleep(delay)
            
            # Calculate score
            if model_error is None:
                score = evaluator.compute_score(entity_responses)
                print(f"   ✅ Total latency: {total_latency:.2f}s | Score: {score:.2%}")
            else:
                score = 0.0
                print(f"   ❌ Failed: {model_error}")
            
            result = TestResult(
                model_name=model,
                entity_responses=entity_responses,
                ground_truth=ground_truth,
                score=score,
                total_latency_seconds=total_latency,
                error=model_error
            )
            
            results.append(result)
            
            # No delay between models - proceed immediately
            # The response time itself provides natural spacing
        
        # Generate Excel report
        print("\n" + "=" * 70)
        print("📊 Generating Excel report...")
        output_path = self.reporter.generate(results)
        
        # Print summary
        self._print_summary(results)
        
        self.client.close()
        
        return results
    
    def _load_models(self) -> List[str]:
        """Load model list from CSV file."""
        csv_path = self.config.project_root / self.config.model_list_file
        
        with open(csv_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        return ModelListParser.parse_model_list(content)
    
    def _print_summary(self, results: List[TestResult]):
        """Print summary statistics."""
        print("\n" + "=" * 70)
        print("📈 SUMMARY")
        print("=" * 70)
        
        successful = [r for r in results if r.error is None]
        failed = [r for r in results if r.error is not None]
        
        print(f"\nTotal models tested: {len(results)}")
        print(f"Successful: {len(successful)}")
        print(f"Failed: {len(failed)}")
        
        if successful:
            scores = [r.score for r in successful]
            latencies = [r.total_latency_seconds for r in successful]
            
            print(f"\nScore Statistics:")
            print(f"  Average: {sum(scores)/len(scores):.2%}")
            print(f"  Min: {min(scores):.2%}")
            print(f"  Max: {max(scores):.2%}")
            
            print(f"\nLatency Statistics:")
            print(f"  Average: {sum(latencies)/len(latencies):.2f}s")
            print(f"  Min: {min(latencies):.2f}s")
            print(f"  Max: {max(latencies):.2f}s")
        
        if failed:
            print(f"\nFailed Models:")
            for r in failed:
                print(f"  - {r.model_name}: {r.error}")


# =============================================================================
# Entry Point
# =============================================================================

def main():
    """Main entry point."""
    config = Config.from_env()
    
    print("\n" + "=" * 70)
    print("Configuration:")
    print(f"  Base URL: {config.base_url}")
    print(f"  Max delay between requests: {config.delay_between_requests}s")
    print(f"  Max tokens: {config.max_tokens}")
    print(f"  Temperature: {config.temperature}")
    print(f"  Request timeout: {config.request_timeout}s")
    print("=" * 70)
    
    runner = ModelTestRunner(config)
    results = runner.run()
    
    print("\n✅ Evaluation complete!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
