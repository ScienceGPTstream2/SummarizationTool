#!/usr/bin/env python3
"""
Llama Context Window Quality Test
===================================
Tests Llama 4 Maverick and Scout across progressively larger context windows
using a needle-in-a-haystack approach.

Two extractions per (model × context_size):
  1. Study author(s)  — appears in first ~500 chars, should always succeed
  2. Needle           — 9-digit number injected at the BOTTOM of each context slice

If Llama's advertised context window is functional, the needle should be found
at all context sizes. If the effective context is limited (or the fallback strategy
always truncates to 2000 chars), the needle will disappear above a threshold.

Usage:
    python backend/scripts/llama_context_window_test.py

Output:
    backend/scripts/llama_context_window_results.xlsx
"""

import asyncio
import json
import os
import sys
import time
from pathlib import Path

# ── Bootstrap: add backend/ to sys.path and load secrets ──────────────────────
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))


def load_secrets() -> None:
    try:
        import toml
    except ImportError:
        print("ERROR: 'toml' not installed. Run: pip install toml")
        sys.exit(1)
    secrets_path = ROOT / "backend" / "core" / "secrets.toml"
    if not secrets_path.exists():
        print(f"ERROR: secrets.toml not found at {secrets_path}")
        sys.exit(1)
    secrets = toml.load(secrets_path)
    for section, kv in secrets.items():
        for k, v in kv.items():
            os.environ[f"{section.upper()}_{k.upper()}"] = str(v)


load_secrets()

from services.llm.llama import LlamaLLMClient  # noqa: E402

# ── Constants ──────────────────────────────────────────────────────────────────
STUDY_FILE = (
    ROOT / "62a2a704bf8bb8b08a401a63f4b3b9a9ef6096446bd0d8ac7e0c70a63c40ad4e_base.md"
)
OUTPUT_FILE = ROOT / "backend" / "scripts" / "llama_context_window_results.xlsx"

NEEDLE = "271834956"  # 9-digit number starting with 2
NEEDLE_MARKER = f"\n\nRESEARCH_ID: {NEEDLE}\n"

MODELS = {
    "maverick": "meta/llama-4-maverick-17b-128e-instruct-maas",
    "scout": "meta/llama-4-scout-17b-16e-instruct-maas",
}

# Context sizes in chars (2K → 128K)
CONTEXT_SIZES = [
    2_000,
    4_000,
    8_000,
    16_000,
    32_000,
    48_000,
    64_000,
    80_000,
    96_000,
    128_000,
]

# Extraction definitions: (name, prompt)
EXTRACTIONS = [
    (
        "author",
        "Extract the study author(s).",
    ),
    (
        "needle",
        (
            "Find the 9-digit research identifier number in the document that starts with 2. "
            "It appears near the end of the document labeled as RESEARCH_ID. "
            "Return only the number."
        ),
    ),
]

# ── Context construction ───────────────────────────────────────────────────────


def make_context(full_doc: str, target_chars: int) -> str:
    """Return first target_chars of the document with the needle appended at the end."""
    return full_doc[:target_chars] + NEEDLE_MARKER


# ── Correctness checks ─────────────────────────────────────────────────────────


def is_correct(extraction_name: str, answer) -> bool:
    if answer is None:
        return False
    answer_str = str(answer).lower()
    if extraction_name == "author":
        return "weiner" in answer_str
    if extraction_name == "needle":
        return NEEDLE in answer_str
    return False


# ── Parse result from extract_entities_with_llama ─────────────────────────────


def parse_result(
    result: dict, ext_name: str, ctx_chars: int, elapsed: float, model_label: str
) -> dict:
    success = result.get("success", False)
    answer = result.get("answer")
    error = result.get("error", "")

    # Token counts: primary strategy stores in meta, fallback stores in raw.usage
    meta = result.get("meta", {}) or {}
    raw = result.get("raw") or {}
    usage = raw.get("usage", {}) if isinstance(raw, dict) else {}
    prompt_tokens = meta.get("prompt_tokens") or usage.get("prompt_tokens")
    completion_tokens = meta.get("completion_tokens") or usage.get("completion_tokens")

    # Strategy detection: more reliable than timing heuristic
    strategy = meta.get("strategy", "unknown")
    used_fallback = strategy == "fallback_minimal"
    # Also flag as likely fallback if response was very fast for large contexts
    # (primary strategy at 16K+ chars takes >30s due to prefill cost)
    timing_fallback = (elapsed < 10.0) and (ctx_chars > 4_000)
    likely_fallback = used_fallback or timing_fallback

    correct = is_correct(ext_name, answer) if success else False

    answer_str = ""
    if answer is not None:
        if isinstance(answer, (dict, list)):
            answer_str = json.dumps(answer, ensure_ascii=False)
        else:
            answer_str = str(answer)
    # Truncate very long answers for Excel readability
    if len(answer_str) > 500:
        answer_str = answer_str[:497] + "..."

    return {
        "Model": model_label,
        "Context Chars": ctx_chars,
        "Context Tokens (est.)": ctx_chars // 3,
        "Extraction": ext_name,
        "Elapsed (s)": round(elapsed, 2),
        "Success": success,
        "Answer": answer_str,
        "Correct": correct,
        "Strategy": strategy,
        "Likely Used Fallback": likely_fallback,
        "Prompt Tokens": prompt_tokens,
        "Completion Tokens": completion_tokens,
        "Error": str(error)[:300] if error else "",
    }


# ── Main test runner ───────────────────────────────────────────────────────────


async def run_model_tests(
    client: LlamaLLMClient, full_doc: str, model_id: str, model_label: str
) -> list[dict]:
    rows = []
    for ctx_chars in CONTEXT_SIZES:
        context = make_context(full_doc, ctx_chars)
        for ext_name, ext_prompt in EXTRACTIONS:
            print(
                f"  [{model_label}] ctx={ctx_chars:>7,} chars  "
                f"({ctx_chars // 3:>5,} tokens est.)  "
                f"extraction={ext_name} ...",
                flush=True,
            )
            t0 = time.perf_counter()
            try:
                result = await client.extract_entities_with_llama(
                    context,
                    ext_prompt,
                    model_name=model_id,
                    max_tokens=512,
                    max_input_length=999_999,  # Disable internal truncation
                )
            except Exception as exc:
                result = {
                    "success": False,
                    "error": f"Exception: {exc}",
                    "answer": None,
                    "meta": {},
                    "raw": None,
                }
            elapsed = time.perf_counter() - t0

            row = parse_result(result, ext_name, ctx_chars, elapsed, model_label)
            rows.append(row)

            status = "CORRECT  " if row["Correct"] else "WRONG    "
            fb = " [FALLBACK]" if row["Likely Used Fallback"] else ""
            ans_preview = (row["Answer"] or "")[:80].replace("\n", " ")
            print(
                f"    -> {elapsed:6.1f}s  {status}{fb}  strategy={row['Strategy']}  "
                f"answer={ans_preview}",
                flush=True,
            )
    return rows


async def main() -> None:
    if not STUDY_FILE.exists():
        print(f"ERROR: Study file not found: {STUDY_FILE}")
        sys.exit(1)

    full_doc = STUDY_FILE.read_text(encoding="utf-8")
    print(f"Study document loaded: {len(full_doc):,} chars")
    print(f"Context sizes: {CONTEXT_SIZES}")
    print(f"Models: {list(MODELS.keys())}")
    print(f"Needle: '{NEEDLE}' injected at bottom of each context slice\n")

    client = LlamaLLMClient()
    if client.disabled:
        print(
            "ERROR: LlamaLLMClient is disabled (missing credentials). Check secrets.toml."
        )
        sys.exit(1)

    all_rows: list[dict] = []
    for model_label, model_id in MODELS.items():
        print(f"\n{'=' * 70}")
        print(f"Model: {model_label}  ({model_id})")
        print(f"{'=' * 70}")
        rows = await run_model_tests(client, full_doc, model_id, model_label)
        all_rows.extend(rows)

    write_excel(all_rows)
    print(f"\nExcel written to: {OUTPUT_FILE}")
    print_summary(all_rows)


# ── Excel output ───────────────────────────────────────────────────────────────


def write_excel(rows: list[dict]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed. Run: pip install openpyxl")
        print("Falling back to CSV output...")
        write_csv_fallback(rows)
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Results ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Results"

    if not rows:
        ws["A1"] = "No results recorded."
        wb.save(OUTPUT_FILE)
        return

    headers = list(rows[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    correct_col = headers.index("Correct") + 1
    fallback_col = headers.index("Likely Used Fallback") + 1

    for row in rows:
        ws.append([row[h] for h in headers])
        row_idx = ws.max_row
        # Colour Correct column
        correct_cell = ws.cell(row=row_idx, column=correct_col)
        if correct_cell.value is True:
            correct_cell.fill = green_fill
        elif correct_cell.value is False:
            correct_cell.fill = red_fill
        # Colour fallback column
        fb_cell = ws.cell(row=row_idx, column=fallback_col)
        if fb_cell.value is True:
            fb_cell.fill = yellow_fill

    # Auto-width
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, min(val_len, 60))
        ws.column_dimensions[col_letter].width = max_len + 3

    # ── Sheet 2: Summary ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary_headers = [
        "Model",
        "Context Chars",
        "Context Tokens (est.)",
        "Extraction",
        "Elapsed (s)",
        "Success",
        "Correct",
        "Likely Used Fallback",
        "Strategy",
    ]
    ws2.append(summary_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws2.append([row.get(h) for h in summary_headers])
        row_idx = ws2.max_row
        correct_col2 = summary_headers.index("Correct") + 1
        fb_col2 = summary_headers.index("Likely Used Fallback") + 1
        c = ws2.cell(row=row_idx, column=correct_col2)
        if c.value is True:
            c.fill = green_fill
        elif c.value is False:
            c.fill = red_fill
        f = ws2.cell(row=row_idx, column=fb_col2)
        if f.value is True:
            f.fill = yellow_fill

    for col_idx, header in enumerate(summary_headers, 1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = max(len(header), 12) + 3

    wb.save(OUTPUT_FILE)


def write_csv_fallback(rows: list[dict]) -> None:
    import csv

    csv_path = OUTPUT_FILE.with_suffix(".csv")
    if not rows:
        return
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"CSV written to: {csv_path}")


# ── Console summary ────────────────────────────────────────────────────────────


def print_summary(rows: list[dict]) -> None:
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(
        f"{'Model':<12} {'Ctx Chars':>10} {'Extraction':<10} {'Elapsed':>8}  {'Correct':<8} {'Fallback'}"
    )
    print("-" * 70)
    for row in rows:
        correct_str = "YES" if row["Correct"] else "NO "
        fallback_str = "YES" if row["Likely Used Fallback"] else "no "
        print(
            f"{row['Model']:<12} {row['Context Chars']:>10,} {row['Extraction']:<10} "
            f"{row['Elapsed (s)']:>7.1f}s  {correct_str:<8} {fallback_str}"
        )


if __name__ == "__main__":
    asyncio.run(main())
